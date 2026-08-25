"""
Transport Cards Accounting System - Web Application
Flask-based web interface with JSON storage (easily migrable to SQL).
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, make_response
from functools import wraps
import os
import json
import uuid
import shutil
import zipfile
from datetime import datetime
from io import BytesIO

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

from app.storage import load_all, save_all, insert, update, delete, get_next_number
from app.models import (
    get_cards, get_card_by_number, create_or_update_card,
    get_card_types, get_card_type_by_id, get_card_type_by_name,
    get_owners, get_owner_by_id, create_owner_if_not_exists,
    get_applicants, get_applicant_by_id, create_applicant_if_not_exists,
    get_organizations, get_organization_by_id,
    get_mfcs, get_mfc_by_id,
    get_employees, get_employee_by_id, get_employee_by_login, check_permission,
    get_documents, get_document_by_id, post_document, delete_document,
    get_cards_report_as_of, get_period_report, get_period_report_detail, get_edo_report, get_summary_report, get_stock_report, get_cards_as_of_report,
    CARD_STATUSES, DOCUMENT_TYPES, REPORT_STATUSES, log_action, now_iso
)

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = "transport_cards_secret_key_change_in_production"
app.config['PERMANENT_SESSION_LIFETIME'] = 1800  # 30 minutes session timeout

# Store for active sessions: {user_id: last_activity_timestamp}
active_sessions = {}

# Document number prefixes by type
DOC_PREFIXES = {
    "receipt": "ПР",
    "print": "Печ",
    "issue": "Выд",
    "transfer_mfc": "ПМФЦ",
    "return_mfc": "ВМФЦ",
    "defect": "Брак",
    "transfer_region": "ПРег"
}

# ============== AUTH DECORATORS ==============
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Необходима авторизация", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Необходима авторизация", "warning")
            return redirect(url_for("login"))
        user = get_employee_by_id(session["user_id"])
        if not user or "admin" not in user.get("roles", []):
            flash("Доступ запрещен", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


def is_issue_user():
    """Check if current user has 'issue' role only."""
    if "user_id" not in session:
        return False
    user = get_employee_by_id(session["user_id"])
    if not user:
        return False
    roles = user.get("roles", [])
    return "issue" in roles and "admin" not in roles and "user" not in roles


def is_reports_user():
    """Check if current user has 'reports' role only (access to reports only)."""
    if "user_id" not in session:
        return False
    user = get_employee_by_id(session["user_id"])
    if not user:
        return False
    roles = user.get("roles", [])
    # User with 'reports' role but no 'admin', 'user', or 'issue' roles
    return "reports" in roles and "admin" not in roles and "user" not in roles and "issue" not in roles


# ============== CONTEXT PROCESSOR ==============
@app.context_processor
def inject_globals():
    user = None
    if "user_id" in session:
        user = get_employee_by_id(session["user_id"])
        # Update last activity time
        active_sessions[user["id"]] = datetime.now()
    
    # Get list of currently active employees (worked in the last 35 minutes)
    from datetime import timedelta
    active_threshold = datetime.now() - timedelta(minutes=35)
    active_employee_ids = [uid for uid, ts in active_sessions.items() if ts >= active_threshold]
    active_employees = [get_employee_by_id(uid) for uid in active_employee_ids if get_employee_by_id(uid)]
    
    return {
        "card_statuses": CARD_STATUSES,
        "document_types": DOCUMENT_TYPES,
        "CARD_STATUSES": CARD_STATUSES,
        "current_user": user,
        "is_admin": user and "admin" in user.get("roles", []),
        "is_issue_user": is_issue_user(),
        "is_reports_user": is_reports_user(),
        "active_employees": active_employees
    }


# ============== MAIN ROUTES ==============
@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_name = request.form.get("login", "").strip()
        password = request.form.get("password", "").strip()
        user = get_employee_by_login(login_name)
        if user and user.get("password") == password:
            session["user_id"] = user["id"]
            session["user_name"] = user.get("full_name", "")
            session.permanent = True
            active_sessions[user["id"]] = datetime.now()
            log_action(user["id"], "LOGIN", f"User {login_name} logged in")
            flash(f"Добро пожаловать, {user.get('full_name', '')}!", "success")
            return redirect(url_for("index"))
        flash("Неверный логин или пароль", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    if "user_id" in session:
        user_id = session["user_id"]
        log_action(user_id, "LOGOUT", "User logged out")
        # Remove from active sessions
        if user_id in active_sessions:
            del active_sessions[user_id]
    session.clear()
    flash("Вы вышли из системы", "info")
    return redirect(url_for("login"))


@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    user = get_employee_by_id(session["user_id"])
    target_id = request.args.get("user_id") or request.form.get("user_id")
    # Admin can change any password; regular users only their own
    if target_id and target_id != session["user_id"]:
        if not user or "admin" not in user.get("roles", []):
            flash("Доступ запрещен", "danger")
            return redirect(url_for("index"))
        target_user = get_employee_by_id(target_id)
    else:
        target_user = user
        target_id = session["user_id"]

    if not target_user:
        flash("Пользователь не найден", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        old_password = request.form.get("old_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        # Non-admin must provide correct old password
        if "admin" not in user.get("roles", []):
            if target_user.get("password") != old_password:
                flash("Неверный текущий пароль", "danger")
                return redirect(url_for("change_password", user_id=target_id))

        if not new_password:
            flash("Введите новый пароль", "danger")
            return redirect(url_for("change_password", user_id=target_id))

        if new_password != confirm_password:
            flash("Пароли не совпадают", "danger")
            return redirect(url_for("change_password", user_id=target_id))

        update("employees", lambda e: e.get("id") == target_id, {"password": new_password})
        log_action(session["user_id"], "CHANGE_PASSWORD", f"Changed password for {target_user.get('login', '')}")
        flash("Пароль изменен", "success")
        return redirect(url_for("index"))

    return render_template("change_password.html", target_user=target_user, is_admin="admin" in user.get("roles", []))


# ============== REFERENCE: CARDS ==============
@app.route("/refs/cards", methods=["GET", "POST"])
@login_required
def ref_cards():
    if request.method == "POST" and "card_number" in request.form:
        card_number = request.form.get("card_number", "").strip()
        if not card_number or len(card_number) != 19 or not card_number.isdigit():
            flash("Номер карты должен содержать ровно 19 цифр", "danger")
            return redirect(url_for("ref_cards"))
        if get_card_by_number(card_number):
            flash("Карта с таким номером уже существует", "danger")
            return redirect(url_for("ref_cards"))
        insert("cards", {
            "card_number": card_number,
            "card_type_id": request.form.get("card_type_id", ""),
            "status": request.form.get("status", ""),
            "owner_id": "",
            "applicant_id": ""
        })
        flash("Карта добавлена", "success")
        return redirect(url_for("ref_cards"))

    search_number = request.args.get("search_number", "")
    search_type = request.args.get("search_type", "")
    sort_by = request.args.get("sort_by", "")
    filters = {}
    if search_number:
        filters["card_number"] = search_number
    cards = get_cards(filters=filters, sort_by=sort_by if sort_by else None)
    card_types = {ct["id"]: ct for ct in get_card_types()}
    owners = {o["id"]: o for o in get_owners()}
    applicants = {a["id"]: a for a in get_applicants()}

    if search_type:
        cards = [c for c in cards if c.get("card_type_id") == search_type]

    return render_template("refs/cards.html",
                           cards=cards,
                           card_types=card_types,
                           owners=owners,
                           applicants=applicants,
                           search_number=search_number,
                           search_type=search_type,
                           sort_by=sort_by)


@app.route("/refs/cards/delete/<card_id>", methods=["POST"])
@login_required
def ref_cards_delete(card_id):
    delete("cards", lambda x: x.get("id") == card_id)
    flash("Карта удалена", "success")
    return redirect(url_for("ref_cards"))


# ============== REFERENCE: CARD TYPES ==============
@app.route("/refs/card_types", methods=["GET", "POST"])
@login_required
def ref_card_types():
    if request.method == "POST" and "name" in request.form:
        insert("card_types", {
            "name": request.form.get("name", ""),
            "print_name": request.form.get("print_name", ""),
            "report_name": request.form.get("report_name", "")
        })
        flash("Вид карты добавлен", "success")
        return redirect(url_for("ref_card_types"))
    items = get_card_types()
    return render_template("refs/card_types.html", items=items)


@app.route("/refs/card_types/edit/<item_id>", methods=["GET", "POST"])
@login_required
def ref_card_types_edit(item_id):
    item = get_card_type_by_id(item_id)
    if not item:
        flash("Вид карты не найден", "danger")
        return redirect(url_for("ref_card_types"))
    
    if request.method == "POST":
        update("card_types", lambda x: x.get("id") == item_id, {
            "name": request.form.get("name", ""),
            "print_name": request.form.get("print_name", ""),
            "report_name": request.form.get("report_name", "")
        })
        flash("Вид карты обновлен", "success")
        return redirect(url_for("ref_card_types"))
    
    return render_template("refs/card_types_edit.html", item=item)


@app.route("/refs/card_types/delete/<item_id>", methods=["POST"])
@login_required
def ref_card_types_delete(item_id):
    delete("card_types", lambda x: x.get("id") == item_id)
    flash("Вид карты удален", "success")
    return redirect(url_for("ref_card_types"))


@app.route("/api/card_types", methods=["GET", "POST"])
@login_required
def api_card_types():
    if request.method == "GET":
        return jsonify(get_card_types())
    data = request.get_json() or {}
    item = insert("card_types", {
        "name": data.get("name", ""),
        "print_name": data.get("print_name", ""),
        "report_name": data.get("report_name", "")
    })
    return jsonify(item), 201


# ============== REFERENCE: OWNERS ==============
@app.route("/refs/owners", methods=["GET", "POST"])
@login_required
def ref_owners():
    if request.method == "POST":
        insert("owners", {"full_name": request.form.get("full_name", "")})
        flash("Владелец добавлен", "success")
        return redirect(url_for("ref_owners"))
    items = get_owners()
    applicants = {a["id"]: a for a in get_applicants()}
    return render_template("refs/owners.html", items=items, applicants=applicants)


@app.route("/refs/owners/delete/<item_id>", methods=["POST"])
@login_required
def ref_owners_delete(item_id):
    delete("owners", lambda x: x.get("id") == item_id)
    flash("Владелец удален", "success")
    return redirect(url_for("ref_owners"))


# ============== REFERENCE: APPLICANTS ==============
@app.route("/refs/applicants", methods=["GET", "POST"])
@login_required
def ref_applicants():
    if request.method == "POST":
        insert("applicants", {"full_name": request.form.get("full_name", "")})
        flash("Заявитель добавлен", "success")
        return redirect(url_for("ref_applicants"))
    items = get_applicants()
    return render_template("refs/applicants.html", items=items)


@app.route("/refs/applicants/delete/<item_id>", methods=["POST"])
@login_required
def ref_applicants_delete(item_id):
    delete("applicants", lambda x: x.get("id") == item_id)
    flash("Заявитель удален", "success")
    return redirect(url_for("ref_applicants"))


# ============== REFERENCE: ORGANIZATIONS ==============
@app.route("/refs/organizations", methods=["GET", "POST"])
@login_required
def ref_organizations():
    if request.method == "POST" and "name" in request.form:
        insert("organizations", {"name": request.form.get("name", "")})
        flash("Организация добавлена", "success")
        return redirect(url_for("ref_organizations"))
    items = get_organizations()
    return render_template("refs/organizations.html", items=items)


@app.route("/refs/organizations/delete/<item_id>", methods=["POST"])
@login_required
def ref_organizations_delete(item_id):
    delete("organizations", lambda x: x.get("id") == item_id)
    flash("Организация удалена", "success")
    return redirect(url_for("ref_organizations"))


@app.route("/api/organizations", methods=["GET", "POST"])
@login_required
def api_organizations():
    if request.method == "GET":
        return jsonify(get_organizations())
    data = request.get_json() or {}
    item = insert("organizations", {"name": data.get("name", "")})
    return jsonify(item), 201


# ============== REFERENCE: MFC ==============
@app.route("/refs/mfcs", methods=["GET", "POST"])
@login_required
def ref_mfcs():
    if request.method == "POST" and "code" in request.form:
        insert("mfcs", {
            "code": request.form.get("code", ""),
            "name": request.form.get("name", "")
        })
        flash("МФЦ добавлен", "success")
        return redirect(url_for("ref_mfcs"))
    items = get_mfcs()
    return render_template("refs/mfcs.html", items=items)


@app.route("/refs/mfcs/edit/<item_id>", methods=["GET", "POST"])
@login_required
def ref_mfcs_edit(item_id):
    item = get_mfc_by_id(item_id)
    if not item:
        flash("МФЦ не найден", "danger")
        return redirect(url_for("ref_mfcs"))
    
    if request.method == "POST":
        update("mfcs", lambda x: x.get("id") == item_id, {
            "code": request.form.get("code", ""),
            "name": request.form.get("name", "")
        })
        flash("МФЦ обновлен", "success")
        return redirect(url_for("ref_mfcs"))
    
    return render_template("refs/mfcs_edit.html", item=item)


@app.route("/refs/mfcs/delete/<item_id>", methods=["POST"])
@login_required
def ref_mfcs_delete(item_id):
    delete("mfcs", lambda x: x.get("id") == item_id)
    flash("МФЦ удален", "success")
    return redirect(url_for("ref_mfcs"))


@app.route("/api/mfcs", methods=["GET", "POST"])
@login_required
def api_mfcs():
    if request.method == "GET":
        return jsonify(get_mfcs())
    data = request.get_json() or {}
    item = insert("mfcs", {
        "code": data.get("code", ""),
        "name": data.get("name", "")
    })
    return jsonify(item), 201


# ============== REFERENCE: EMPLOYEES ==============
@app.route("/refs/employees", methods=["GET", "POST"])
@login_required
@admin_required
def ref_employees():
    if request.method == "POST":
        roles = request.form.getlist("roles")
        insert("employees", {
            "full_name": request.form.get("full_name", ""),
            "login": request.form.get("login", ""),
            "password": request.form.get("password", ""),
            "roles": roles if roles else ["user"],
            "permissions": {}
        })
        flash("Сотрудник добавлен", "success")
        return redirect(url_for("ref_employees"))
    items = get_employees()
    return render_template("refs/employees.html", items=items)


@app.route("/refs/employees/edit/<item_id>", methods=["GET", "POST"])
@login_required
@admin_required
def ref_employees_edit(item_id):
    item = get_employee_by_id(item_id)
    if not item:
        flash("Сотрудник не найден", "danger")
        return redirect(url_for("ref_employees"))
    
    if request.method == "POST":
        roles = request.form.getlist("roles")
        update("employees", lambda x: x.get("id") == item_id, {
            "full_name": request.form.get("full_name", ""),
            "login": request.form.get("login", ""),
            "roles": roles if roles else ["user"],
            "permissions": {}
        })
        flash("Сотрудник обновлен", "success")
        return redirect(url_for("ref_employees"))
    
    return render_template("refs/employees_edit.html", item=item)


@app.route("/refs/employees/delete/<item_id>", methods=["POST"])
@login_required
@admin_required
def ref_employees_delete(item_id):
    delete("employees", lambda x: x.get("id") == item_id)
    flash("Сотрудник удален", "success")
    return redirect(url_for("ref_employees"))


@app.route("/api/employees", methods=["GET", "POST"])
@login_required
@admin_required
def api_employees():
    if request.method == "GET":
        return jsonify(get_employees())
    data = request.get_json() or {}
    item = insert("employees", {
        "full_name": data.get("full_name", ""),
        "login": data.get("login", ""),
        "password": data.get("password", ""),
        "roles": data.get("roles", []),
        "permissions": data.get("permissions", {})
    })
    return jsonify(item), 201


# ============== DOCUMENTS ==============
@app.route("/docs")
@login_required
def docs_journal():
    doc_type = request.args.get("type", "")
    if is_issue_user():
        # Issue users can only see issue documents
        if doc_type and doc_type != "issue":
            flash("Доступ запрещен", "danger")
            return redirect(url_for("docs_journal"))
        items = get_documents(doc_type="issue")
        doc_type = "issue"
    else:
        items = get_documents(doc_type=doc_type if doc_type else None)
    employees = get_employees()
    emp_map = {e["id"]: e for e in employees}
    return render_template("docs/journal.html", items=items, doc_type=doc_type, emp_map=emp_map)


@app.route("/docs/create/<doc_type>", methods=["GET", "POST"])
@login_required
def doc_create(doc_type):
    if doc_type not in DOCUMENT_TYPES:
        flash("Неизвестный тип документа", "danger")
        return redirect(url_for("docs_journal"))
    if is_issue_user() and doc_type != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))

    if request.method == "POST" and "doc_date" in request.form:
        lines = []
        idx = 1
        while True:
            card_num = request.form.get(f"line_{idx}_card_number", "")
            if not card_num:
                break
            lines.append({
                "line_no": idx,
                "card_number": card_num,
                "card_type_id": request.form.get(f"line_{idx}_card_type_id", ""),
                "owner_name": request.form.get(f"line_{idx}_owner_name", ""),
                "applicant_name": request.form.get(f"line_{idx}_applicant_name", "")
            })
            idx += 1

        doc = insert("documents", {
            "doc_type": doc_type,
            "doc_number": request.form.get("doc_number") or get_next_number(DOC_PREFIXES.get(doc_type, "DOC")),
            "doc_date": request.form.get("doc_date", datetime.now().strftime("%Y-%m-%d")),
            "organization_id": request.form.get("organization_id", ""),
            "mfc_id": request.form.get("mfc_id", ""),
            "employee_id": request.form.get("employee_id", ""),
            "lines": lines,
            "status": "draft",
            "created_by": session.get("user_id"),
            "created_at": now_iso()
        })
        flash("Документ создан", "success")
        return redirect(url_for("doc_edit", doc_id=doc["id"]))

    # GET
    card_types = get_card_types()
    organizations = get_organizations()
    mfcs = get_mfcs()
    employees = get_employees()
    return render_template("docs/create.html",
                           doc_type=doc_type,
                           doc_type_name=DOCUMENT_TYPES[doc_type],
                           card_types=card_types,
                           organizations=organizations,
                           mfcs=mfcs,
                           employees=employees,
                           current_date=datetime.now().strftime("%Y-%m-%d"))


@app.route("/docs/edit/<doc_id>", methods=["GET", "POST"])
@login_required
def doc_edit(doc_id):
    doc = get_document_by_id(doc_id)
    if not doc:
        flash("Документ не найден", "danger")
        return redirect(url_for("docs_journal"))
    if is_issue_user() and doc.get("doc_type") != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))

    if request.method == "POST":
        lines = []
        idx = 1
        while True:
            card_num = request.form.get(f"line_{idx}_card_number", "")
            if not card_num:
                break
            lines.append({
                "line_no": idx,
                "card_number": card_num,
                "card_type_id": request.form.get(f"line_{idx}_card_type_id", ""),
                "owner_name": request.form.get(f"line_{idx}_owner_name", ""),
                "applicant_name": request.form.get(f"line_{idx}_applicant_name", "")
            })
            idx += 1

        update("documents", lambda d: d.get("id") == doc_id, {
            "doc_number": request.form.get("doc_number", doc.get("doc_number")),
            "doc_date": request.form.get("doc_date", doc.get("doc_date")),
            "organization_id": request.form.get("organization_id", doc.get("organization_id")),
            "mfc_id": request.form.get("mfc_id", doc.get("mfc_id")),
            "employee_id": request.form.get("employee_id", doc.get("employee_id")),
            "lines": lines,
            "updated_at": now_iso()
        })
        flash("Документ сохранен", "success")
        return redirect(url_for("doc_edit", doc_id=doc_id))

    card_types = get_card_types()
    ct_map = {ct["id"]: ct for ct in card_types}
    organizations = get_organizations()
    mfcs = get_mfcs()
    employees = get_employees()
    cards = get_cards()
    author = get_employee_by_id(doc.get("created_by")) if doc.get("created_by") else None
    return render_template("docs/edit.html",
                           doc=doc,
                           doc_type_name=DOCUMENT_TYPES.get(doc.get("doc_type"), ""),
                           card_types=card_types,
                           ct_map=ct_map,
                           organizations=organizations,
                           mfcs=mfcs,
                           employees=employees,
                           cards=cards,
                           author=author)


@app.route("/docs/post/<doc_id>", methods=["POST"])
@login_required
def doc_post_route(doc_id):
    doc = get_document_by_id(doc_id)
    if is_issue_user() and (not doc or doc.get("doc_type") != "issue"):
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))
    success, message = post_document(doc_id, session.get("user_id"))
    if success:
        flash(message, "success")
    else:
        flash(message, "danger")
    return redirect(url_for("doc_edit", doc_id=doc_id))


@app.route("/docs/delete/<doc_id>", methods=["POST"])
@login_required
def doc_delete_route(doc_id):
    doc = get_document_by_id(doc_id)
    if is_issue_user() and (not doc or doc.get("doc_type") != "issue"):
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))
    if delete_document(doc_id, session.get("user_id")):
        flash("Документ удален", "success")
    else:
        flash("Ошибка удаления документа", "danger")
    return redirect(url_for("docs_journal"))


# ============== PRINT FORMS ==============
@app.route("/docs/print/<doc_id>/<form_type>")
@login_required
def doc_print_form(doc_id, form_type):
    doc = get_document_by_id(doc_id)
    if not doc:
        flash("Документ не найден", "danger")
        return redirect(url_for("docs_journal"))
    if is_issue_user() and doc.get("doc_type") != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))

    ct_map = {ct["id"]: ct for ct in get_card_types()}
    org = get_organization_by_id(doc.get("organization_id", "")) if doc.get("organization_id") else None
    mfc = get_mfc_by_id(doc.get("mfc_id", "")) if doc.get("mfc_id") else None
    employee = get_employee_by_id(doc.get("employee_id", "")) if doc.get("employee_id") else None
    const_list = load_all("constants")
    our_org = const_list[0].get("organization_name", "ООО Транспортные Карты") if const_list else "ООО Транспортные Карты"
    author = get_employee_by_id(doc.get("created_by")) if doc.get("created_by") else None

    # Enrich lines with card data
    enriched_lines = []
    for line in doc.get("lines", []):
        card = get_card_by_number(line.get("card_number", ""))
        ct = ct_map.get(line.get("card_type_id", ""), {}) if line.get("card_type_id") else (ct_map.get(card.get("card_type_id", ""), {}) if card else {})
        enriched_lines.append({
            "line_no": line.get("line_no"),
            "card_number": line.get("card_number"),
            "card_type_name": ct.get("name", ""),
            "card_type_print_name": ct.get("print_name", ct.get("name", "")),
            "owner_name": line.get("owner_name", ""),
            "applicant_name": line.get("applicant_name", "")
        })

    if doc.get("doc_type") == "return_mfc" and form_type == "act":
        return render_template("docs/print_act.html",
                               doc=doc,
                               lines=enriched_lines,
                               mfc=mfc,
                               employee=employee,
                               author=author,
                               our_org=our_org)
    elif doc.get("doc_type") == "return_mfc" and form_type == "register":
        return render_template("docs/print_register.html",
                               doc=doc,
                               lines=enriched_lines,
                               mfc=mfc,
                               employee=employee,
                               author=author,
                               our_org=our_org)
    elif doc.get("doc_type") == "transfer_mfc" and form_type == "register":
        return render_template("docs/print_register.html",
                               doc=doc,
                               lines=enriched_lines,
                               mfc=mfc,
                               employee=employee,
                               author=author,
                               our_org=our_org)
    elif doc.get("doc_type") == "transfer_mfc" and form_type == "transfer":
        return render_template("docs/print_transfer_mfc.html",
                               doc=doc,
                               lines=enriched_lines,
                               mfc=mfc,
                               employee=employee,
                               author=author,
                               our_org=our_org)
    else:
        # Generic print form for any document
        return render_template("docs/print_generic.html",
                               doc=doc,
                               doc_type_name=DOCUMENT_TYPES.get(doc.get("doc_type"), ""),
                               lines=enriched_lines,
                               org=org,
                               mfc=mfc,
                               employee=employee,
                               author=author,
                               our_org=our_org)


# ============== API: CARDS FOR PICKING ==============
@app.route("/api/cards")
@login_required
def api_cards():
    status = request.args.get("status", "")
    q = request.args.get("q", "")
    owner_q = request.args.get("owner", "")
    cards = get_cards()
    owners = {o["id"]: o for o in get_owners()}
    if status:
        cards = [c for c in cards if c.get("status") == status]
    if q:
        cards = [c for c in cards if q in c.get("card_number", "")]
    if owner_q:
        # Filter by owner full name (case-insensitive substring)
        cards = [c for c in cards
                 if owner_q.lower() in owners.get(c.get("owner_id", ""), {}).get("full_name", "").lower()]
    result = []
    ct_map = {ct["id"]: ct for ct in get_card_types()}
    for c in cards:
        ct = ct_map.get(c.get("card_type_id", ""), {})
        owner = owners.get(c.get("owner_id", ""), {})
        result.append({
            "id": c["id"],
            "card_number": c.get("card_number"),
            "card_type_id": c.get("card_type_id"),
            "card_type_name": ct.get("name", ""),
            "status": c.get("status"),
            "status_name": CARD_STATUSES.get(c.get("status"), ""),
            "owner_id": c.get("owner_id"),
            "owner_name": owner.get("full_name", ""),
            "applicant_id": c.get("applicant_id")
        })
    return jsonify(result)


@app.route("/api/card/<number>")
@login_required
def api_card_by_number(number):
    card = get_card_by_number(number)
    if not card:
        return jsonify({"error": "Card not found"}), 404
    ct = get_card_type_by_id(card.get("card_type_id", ""))
    owner = get_owner_by_id(card.get("owner_id", ""))
    return jsonify({
        "id": card["id"],
        "card_number": card.get("card_number"),
        "card_type_id": card.get("card_type_id"),
        "card_type_name": ct.get("name", "") if ct else "",
        "status": card.get("status"),
        "status_name": CARD_STATUSES.get(card.get("status"), ""),
        "owner_id": card.get("owner_id"),
        "owner_name": owner.get("full_name", "") if owner else "",
        "applicant_id": card.get("applicant_id")
    })


# ============== EXCEL UPLOAD FOR DOCUMENTS ==============
def _parse_excel_lines(file_storage, doc_type):
    """Parse uploaded Excel file and return lines for document."""
    try:
        import openpyxl
    except ImportError:
        flash("Для работы с Excel необходимо установить openpyxl: pip install openpyxl", "danger")
        return None

    wb = openpyxl.load_workbook(file_storage)
    ws = wb.active
    lines = []
    row_idx = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_idx += 1
        if not row or not row[0]:
            continue
        card_number = str(row[0]).strip()
        if not card_number:
            continue

        line = {
            "line_no": len(lines) + 1,
            "card_number": card_number,
            "card_type_id": "",
            "owner_name": "",
            "applicant_name": ""
        }

        # Auto-fill card_type_id from cards reference for all doc types except receipt
        if doc_type != "receipt":
            card = get_card_by_number(card_number)
            if card:
                line["card_type_id"] = card.get("card_type_id", "")

        if doc_type == "receipt":
            if len(row) > 1 and row[1]:
                ct = get_card_type_by_name(str(row[1]).strip())
                if ct:
                    line["card_type_id"] = ct["id"]
        elif doc_type == "print":
            if len(row) > 1 and row[1]:
                line["owner_name"] = str(row[1]).strip()
        elif doc_type == "defect":
            pass
        elif doc_type == "transfer_region":
            pass
        elif doc_type in ["issue", "transfer_mfc", "return_mfc"]:
            if len(row) > 1 and row[1]:
                line["owner_name"] = str(row[1]).strip()
            if len(row) > 2 and row[2]:
                line["applicant_name"] = str(row[2]).strip()

        lines.append(line)

    return lines


@app.route("/docs/upload_excel/<doc_type>", methods=["POST"])
@login_required
def doc_upload_excel(doc_type):
    if doc_type not in DOCUMENT_TYPES:
        flash("Неизвестный тип документа", "danger")
        return redirect(url_for("docs_journal"))
    if is_issue_user() and doc_type != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))

    file = request.files.get("excel_file")
    if not file:
        flash("Файл не выбран", "danger")
        return redirect(url_for("doc_create", doc_type=doc_type))

    lines = _parse_excel_lines(file, doc_type)
    if lines is None:
        return redirect(url_for("doc_create", doc_type=doc_type))

    doc = insert("documents", {
        "doc_type": doc_type,
        "doc_number": get_next_number(DOC_PREFIXES.get(doc_type, "DOC")),
        "doc_date": datetime.now().strftime("%Y-%m-%d"),
        "organization_id": request.form.get("organization_id", ""),
        "mfc_id": request.form.get("mfc_id", ""),
        "employee_id": request.form.get("employee_id", ""),
        "lines": lines,
        "status": "draft",
        "created_by": session.get("user_id"),
        "created_at": now_iso()
    })
    flash(f"Документ создан, загружено {len(lines)} строк из Excel", "success")
    return redirect(url_for("doc_edit", doc_id=doc["id"]))


@app.route("/docs/upload_excel_existing/<doc_id>", methods=["POST"])
@login_required
def doc_upload_excel_existing(doc_id):
    doc = get_document_by_id(doc_id)
    if not doc or doc.get("status") == "posted":
        flash("Документ не найден или уже проведен", "danger")
        return redirect(url_for("docs_journal"))
    if is_issue_user() and doc.get("doc_type") != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))

    file = request.files.get("excel_file")
    if not file:
        flash("Файл не выбран", "danger")
        return redirect(url_for("doc_edit", doc_id=doc_id))

    lines = _parse_excel_lines(file, doc.get("doc_type"))
    if lines is None:
        return redirect(url_for("doc_edit", doc_id=doc_id))

    update("documents", lambda d: d.get("id") == doc_id, {
        "lines": lines,
        "updated_at": now_iso()
    })
    flash(f"Загружено {len(lines)} строк из Excel", "success")
    return redirect(url_for("doc_edit", doc_id=doc_id))


@app.route("/api/download_doc_template/<doc_type>")
@login_required
def download_doc_template(doc_type):
    """Download Excel template for document import."""
    if is_issue_user() and doc_type != "issue":
        flash("Доступ запрещен", "danger")
        return redirect(url_for("docs_journal"))
    try:
        import openpyxl
    except ImportError:
        flash("Для работы с Excel необходимо установить openpyxl: pip install openpyxl", "danger")
        return redirect(url_for("index"))

    wb = openpyxl.Workbook()
    ws = wb.active

    if doc_type == "receipt":
        ws.append(["Номер карты", "Вид карты"])
    elif doc_type == "print":
        ws.append(["Номер карты", "ФИО владельца"])
    elif doc_type == "defect":
        ws.append(["Номер карты"])
    elif doc_type == "transfer_region":
        ws.append(["Номер карты"])
    elif doc_type in ["issue", "transfer_mfc", "return_mfc"]:
        ws.append(["Номер карты", "ФИО владельца", "ФИО заявителя"])
    else:
        flash("Для данного документа шаблон не предусмотрен", "warning")
        return redirect(url_for("docs_journal"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"template_{doc_type}.xlsx", as_attachment=True)


# ============== BACKUP ==============
@app.route("/backup", methods=["GET", "POST"])
@login_required
@admin_required
def backup_index():
    # Handle schedule settings POST
    if request.method == "POST" and "backup_schedule" in request.form:
        schedule_time = request.form.get("backup_schedule_time", "02:00")
        backup_enabled = request.form.get("backup_enabled") == "on"
        # Store schedule settings in constants.json
        constants = load_all("constants")
        constants_data = {}
        for c in constants:
            constants_data[c.get("key", "")] = c.get("value", "")
        constants_data["backup_schedule_time"] = schedule_time
        constants_data["backup_enabled"] = "true" if backup_enabled else "false"
        # Save back to constants
        save_all("constants", [{"key": k, "value": v} for k, v in constants_data.items()])
        log_action(session.get("user_id"), "BACKUP_SCHEDULE_UPDATE", f"Schedule updated: {schedule_time}, enabled={backup_enabled}")
        flash("Настройки резервного копирования обновлены", "success")
        return redirect(url_for("backup_index"))
    
    if request.method == "POST":
        file = request.files.get("backup_file")
        if not file:
            flash("Файл не выбран", "danger")
            return redirect(url_for("backup_index"))

        temp_path = os.path.join(BACKUP_DIR, "temp_upload.zip")
        file.save(temp_path)

        try:
            with zipfile.ZipFile(temp_path, "r") as zf:
                zf.testzip()
        except zipfile.BadZipFile:
            os.remove(temp_path)
            flash("Неверный формат файла", "danger")
            return redirect(url_for("backup_index"))

        # Create emergency backup before restore
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        emergency = os.path.join(BACKUP_DIR, f"auto_before_restore_{timestamp}.zip")
        with zipfile.ZipFile(emergency, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(DATA_DIR):
                for f in files:
                    fp = os.path.join(root, f)
                    arcname = os.path.relpath(fp, os.path.dirname(DATA_DIR))
                    zf.write(fp, arcname)

        # Restore
        with zipfile.ZipFile(temp_path, "r") as zf:
            zf.extractall(os.path.dirname(DATA_DIR))

        os.remove(temp_path)
        log_action(session.get("user_id"), "BACKUP_RESTORE", "Restored from uploaded file")
        flash("Данные восстановлены из загруженной резервной копии", "success")
        return redirect(url_for("backup_index"))

    backups = []
    for fname in sorted(os.listdir(BACKUP_DIR), reverse=True):
        if fname.endswith(".zip"):
            fpath = os.path.join(BACKUP_DIR, fname)
            size = os.path.getsize(fpath)
            backups.append({
                "name": fname,
                "size": f"{size / 1024:.1f} КБ",
                "date": datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    # Load schedule settings
    constants = load_all("constants")
    schedule_time = "02:00"
    backup_enabled = False
    for c in constants:
        if c.get("key") == "backup_schedule_time":
            schedule_time = c.get("value", "02:00")
        elif c.get("key") == "backup_enabled":
            backup_enabled = c.get("value") == "true"
    
    return render_template("backup.html", backups=backups, schedule_time=schedule_time, backup_enabled=backup_enabled)


@app.route("/backup/create", methods=["POST"])
@login_required
@admin_required
def backup_create():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"backup_{timestamp}.zip"
    backup_path = os.path.join(BACKUP_DIR, backup_name)

    with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(DATA_DIR):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, os.path.dirname(DATA_DIR))
                zf.write(file_path, arcname)

    log_action(session.get("user_id"), "BACKUP_CREATE", f"Created backup {backup_name}")
    flash(f"Резервная копия создана: {backup_name}", "success")
    return redirect(url_for("backup_index"))


@app.route("/backup/download/<name>")
@login_required
@admin_required
def backup_download(name):
    backup_path = os.path.join(BACKUP_DIR, name)
    if not os.path.exists(backup_path):
        flash("Файл не найден", "danger")
        return redirect(url_for("backup_index"))
    return send_file(backup_path, download_name=name, as_attachment=True)


@app.route("/backup/restore/<name>", methods=["POST"])
@login_required
@admin_required
def backup_restore(name):
    backup_path = os.path.join(BACKUP_DIR, name)
    if not os.path.exists(backup_path):
        flash("Файл не найден", "danger")
        return redirect(url_for("backup_index"))

    # Create emergency backup before restore
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    emergency = os.path.join(BACKUP_DIR, f"auto_before_restore_{timestamp}.zip")
    with zipfile.ZipFile(emergency, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(DATA_DIR):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, os.path.dirname(DATA_DIR))
                zf.write(file_path, arcname)

    # Restore
    with zipfile.ZipFile(backup_path, "r") as zf:
        zf.extractall(os.path.dirname(DATA_DIR))

    log_action(session.get("user_id"), "BACKUP_RESTORE", f"Restored from {name}")
    flash(f"Данные восстановлены из резервной копии: {name}", "success")
    return redirect(url_for("backup_index"))


@app.route("/backup/delete/<name>", methods=["POST"])
@login_required
@admin_required
def backup_delete(name):
    backup_path = os.path.join(BACKUP_DIR, name)
    if os.path.exists(backup_path):
        os.remove(backup_path)
        log_action(session.get("user_id"), "BACKUP_DELETE", f"Deleted {name}")
        flash("Резервная копия удалена", "success")
    return redirect(url_for("backup_index"))


# ============== REPORTS ==============
@app.route("/reports")
@login_required
def reports():
    return render_template("reports/index.html")


@app.route("/reports/cards_as_of", methods=["GET", "POST"])
@login_required
def report_cards_as_of():
    report = None
    date_str = ""
    if request.method == "POST":
        date_str = request.form.get("date", "")
        if date_str:
            report = get_cards_report_as_of(date_str)
    return render_template("reports/cards_as_of.html", report=report, date_str=date_str)


@app.route("/reports/period", methods=["GET", "POST"])
@login_required
def report_period():
    report = None
    start_date = ""
    end_date = ""
    total_print = 0
    total_issue = 0
    if request.method == "POST":
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")
        if start_date and end_date:
            report = get_period_report(start_date, end_date)
            for row in report:
                total_print += row.get("print_count", 0)
                total_issue += row.get("issue_count", 0)
    return render_template("reports/period.html", report=report, start_date=start_date, end_date=end_date,
                           total_print=total_print, total_issue=total_issue)


@app.route("/reports/period/print")
@login_required
def print_period():
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    if not start_date or not end_date:
        flash("Укажите период", "warning")
        return redirect(url_for("report_period"))
    report = get_period_report(start_date, end_date)
    total_print = sum(r.get("print_count", 0) for r in report)
    total_issue = sum(r.get("issue_count", 0) for r in report)
    return render_template("reports/print_period.html",
                           report=report,
                           start_date=start_date,
                           end_date=end_date,
                           total_print=total_print,
                           total_issue=total_issue)


@app.route("/reports/edo", methods=["GET", "POST"])
@login_required
def report_edo():
    report = None
    start_date = ""
    end_date = ""
    if request.method == "POST":
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")
        if start_date and end_date:
            report = get_edo_report(start_date, end_date)
    return render_template("reports/edo.html", report=report, start_date=start_date, end_date=end_date)


@app.route("/reports/summary", methods=["GET", "POST"])
@login_required
def report_summary():
    report = None
    start_date = ""
    end_date = ""
    if request.method == "POST":
        start_date = request.form.get("start_date", "")
        end_date = request.form.get("end_date", "")
        if start_date and end_date:
            report = get_summary_report(start_date, end_date)
    return render_template("reports/summary.html", report=report, start_date=start_date, end_date=end_date)


@app.route("/reports/print_summary")
@login_required
def print_summary():
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    if not start_date or not end_date:
        flash("Укажите период", "warning")
        return redirect(url_for("report_summary"))
    report = get_summary_report(start_date, end_date)
    return render_template("reports/print_summary.html", report=report, start_date=start_date, end_date=end_date)


@app.route("/reports/stock")
@login_required
def report_stock():
    report = get_stock_report()
    total_print = sum(r.get("ready_to_print", 0) for r in report)
    total_issue = sum(r.get("ready_to_issue", 0) for r in report)
    return render_template("reports/stock.html", report=report, total_print=total_print, total_issue=total_issue)


@app.route("/reports/cards_as_of")
@login_required
def report_cards_as_of_new():
    """Report: карты на число."""
    report = get_cards_as_of_report()
    return render_template("reports/cards_as_of.html", report=report, statuses=REPORT_STATUSES)


# ============== REPORTS EXPORT TO EXCEL ==============
def _export_report_to_excel(report_data, columns, sheet_title="Отчет"):
    try:
        import openpyxl
    except ImportError:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(columns)
    for row in report_data:
        ws.append([row.get(col, "") for col in columns])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/reports/export/cards_as_of")
@login_required
def export_cards_as_of():
    date_str = request.args.get("date", "")
    if not date_str:
        flash("Укажите дату", "warning")
        return redirect(url_for("report_cards_as_of"))
    report = get_cards_report_as_of(date_str)
    output = _export_report_to_excel(report, ["card_type_name", "status", "count"], "Карты за день")
    if output is None:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("report_cards_as_of"))
    return send_file(output, download_name=f"cards_as_of_{date_str}.xlsx", as_attachment=True)


@app.route("/reports/export/period")
@login_required
def export_period():
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    if not start_date or not end_date:
        flash("Укажите период", "warning")
        return redirect(url_for("report_period"))
    report = get_period_report(start_date, end_date)
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("report_period"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет за период"
    ws.append(["Вид карты", "Печать карт", "Выдача + Передача в МФЦ"])
    total_print = 0
    total_issue = 0
    for row in report:
        ws.append([
            row.get("card_type_name", ""),
            row.get("print_count", 0),
            row.get("issue_count", 0)
        ])
        total_print += row.get("print_count", 0)
        total_issue += row.get("issue_count", 0)
    ws.append(["Итого", total_print, total_issue])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"period_{start_date}_{end_date}.xlsx", as_attachment=True)


@app.route("/reports/export/edo")
@login_required
def export_edo():
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    if not start_date or not end_date:
        flash("Укажите период", "warning")
        return redirect(url_for("report_edo"))
    report = get_edo_report(start_date, end_date)
    # Custom format: card_type_name, numbers
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("report_edo"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЭДО"
    ws.append(["Вид карты", "Номера карт"])
    for row in report:
        ws.append([row.get("card_type_name", ""), row.get("numbers", "")])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"edo_{start_date}_{end_date}.xlsx", as_attachment=True)


@app.route("/reports/export/summary")
@login_required
def export_summary():
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    if not start_date or not end_date:
        flash("Укажите период", "warning")
        return redirect(url_for("report_summary"))
    report = get_summary_report(start_date, end_date)
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("report_summary"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводный"
    ws.append(["Вид карты", "Номера карт"])
    for row in report:
        ws.append([row.get("card_type_name", ""), ", ".join(row.get("numbers", []))])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"summary_{start_date}_{end_date}.xlsx", as_attachment=True)


@app.route("/reports/export/stock")
@login_required
def export_stock():
    report = get_stock_report()
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("report_stock"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остаток карт"
    ws.append(["Вид карт", "Готова к печати", "Готова к выдачи"])
    total_print = 0
    total_issue = 0
    for row in report:
        ws.append([
            row.get("card_type_name", ""),
            row.get("ready_to_print", 0),
            row.get("ready_to_issue", 0)
        ])
        total_print += row.get("ready_to_print", 0)
        total_issue += row.get("ready_to_issue", 0)
    ws.append(["Итого", total_print, total_issue])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="stock_report.xlsx", as_attachment=True)


# ============== ACTION LOG ==============
@app.route("/action_log")
@login_required
@admin_required
def action_log():
    logs = load_all("action_log")
    logs = sorted(logs, key=lambda x: x.get("timestamp", ""), reverse=True)
    employees = {e["id"]: e for e in get_employees()}
    return render_template("action_log.html", logs=logs, employees=employees)


# ============== CONSTANTS ==============
@app.route("/constants", methods=["GET", "POST"])
@login_required
@admin_required
def constants():
    constants_list = load_all("constants")
    const = constants_list[0] if constants_list else {"organization_name": ""}
    if request.method == "POST":
        update("constants", lambda c: c.get("id") == const.get("id"), {
            "organization_name": request.form.get("organization_name", "")
        })
        flash("Константы сохранены", "success")
        return redirect(url_for("constants"))
    return render_template("constants.html", const=const)


# ============== IMPORT/EXPORT HELPERS ==============
@app.route("/api/download_template/<ref_name>")
@login_required
def download_template(ref_name):
    """Download Excel template for reference import."""
    try:
        import openpyxl
    except ImportError:
        flash("Для работы с Excel необходимо установить openpyxl: pip install openpyxl", "danger")
        return redirect(url_for("index"))

    wb = openpyxl.Workbook()
    ws = wb.active

    if ref_name == "card_types":
        ws.append(["Вид карты", "Наименование для печати"])
    elif ref_name == "organizations":
        ws.append(["Наименование организации"])
    elif ref_name == "mfcs":
        ws.append(["Код", "Наименование"])
    elif ref_name == "cards":
        ws.append(["Номер карты", "Вид карты", "ФИО владельца", "ФИО заявителя"])
    elif ref_name == "owners":
        ws.append(["ФИО владельца"])
    elif ref_name == "applicants":
        ws.append(["ФИО заявителя"])
    elif ref_name == "employees":
        ws.append(["ФИО", "Логин", "Пароль", "Роли"])
    else:
        flash("Неизвестный справочник", "danger")
        return redirect(url_for("index"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"template_{ref_name}.xlsx", as_attachment=True)


# ============== REFERENCE IMPORT FROM EXCEL ==============
def _parse_reference_excel(file_storage, ref_name):
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl не установлен"
    wb = openpyxl.load_workbook(file_storage)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if ref_name == "card_types":
            items.append({"name": str(row[0]).strip(), "print_name": str(row[1]).strip() if len(row) > 1 else ""})
        elif ref_name == "organizations":
            items.append({"name": str(row[0]).strip()})
        elif ref_name == "mfcs":
            items.append({"code": str(row[0]).strip(), "name": str(row[1]).strip() if len(row) > 1 else ""})
        elif ref_name == "cards":
            items.append({
                "card_number": str(row[0]).strip(),
                "card_type_name": str(row[1]).strip() if len(row) > 1 else "",
                "owner_name": str(row[2]).strip() if len(row) > 2 else "",
                "applicant_name": str(row[3]).strip() if len(row) > 3 else ""
            })
        elif ref_name == "owners":
            items.append({"full_name": str(row[0]).strip()})
        elif ref_name == "applicants":
            items.append({"full_name": str(row[0]).strip()})
        elif ref_name == "employees":
            roles = [r.strip() for r in str(row[3]).split(",")] if len(row) > 3 and row[3] else ["user"]
            items.append({
                "full_name": str(row[0]).strip(),
                "login": str(row[1]).strip() if len(row) > 1 else "",
                "password": str(row[2]).strip() if len(row) > 2 else "",
                "roles": roles
            })
    return items, None


@app.route("/refs/import/<ref_name>", methods=["POST"])
@login_required
def ref_import_excel(ref_name):
    file = request.files.get("excel_file")
    if not file:
        flash("Файл не выбран", "danger")
        return redirect(url_for(f"ref_{ref_name}"))
    items, error = _parse_reference_excel(file, ref_name)
    if error:
        flash(error, "danger")
        return redirect(url_for(f"ref_{ref_name}"))
    count = 0
    for item in items:
        if ref_name == "card_types":
            insert("card_types", item)
            count += 1
        elif ref_name == "organizations":
            insert("organizations", item)
            count += 1
        elif ref_name == "mfcs":
            insert("mfcs", item)
            count += 1
        elif ref_name == "cards":
            # Validate card number
            cn = item.get("card_number", "")
            if not cn or len(cn) != 19 or not cn.isdigit():
                continue
            if get_card_by_number(cn):
                continue
            ct = get_card_type_by_name(item.get("card_type_name", ""))
            insert("cards", {
                "card_number": cn,
                "card_type_id": ct["id"] if ct else "",
                "status": "",
                "owner_id": "",
                "applicant_id": ""
            })
            count += 1
        elif ref_name == "owners":
            if item.get("full_name"):
                insert("owners", item)
                count += 1
        elif ref_name == "applicants":
            if item.get("full_name"):
                insert("applicants", item)
                count += 1
        elif ref_name == "employees":
            if item.get("full_name") and item.get("login"):
                insert("employees", {
                    "full_name": item["full_name"],
                    "login": item["login"],
                    "password": item.get("password", ""),
                    "roles": item.get("roles", ["user"]),
                    "permissions": {}
                })
                count += 1
    flash(f"Импортировано {count} записей", "success")
    return redirect(url_for(f"ref_{ref_name}"))


# ============== REFERENCE EXPORT TO EXCEL ==============
@app.route("/refs/export/<ref_name>")
@login_required
def ref_export_excel(ref_name):
    try:
        import openpyxl
    except ImportError:
        flash("openpyxl не установлен", "danger")
        return redirect(url_for("index"))
    wb = openpyxl.Workbook()
    ws = wb.active
    if ref_name == "cards":
        ws.append(["Номер карты", "Вид карты", "Статус", "ФИО владельца", "ФИО заявителя"])
        cards = get_cards()
        ct_map = {ct["id"]: ct for ct in get_card_types()}
        owners = {o["id"]: o for o in get_owners()}
        applicants = {a["id"]: a for a in get_applicants()}
        for c in cards:
            ws.append([
                c.get("card_number", ""),
                ct_map.get(c.get("card_type_id", ""), {}).get("name", ""),
                CARD_STATUSES.get(c.get("status", ""), c.get("status", "")),
                owners.get(c.get("owner_id", ""), {}).get("full_name", ""),
                applicants.get(c.get("applicant_id", ""), {}).get("full_name", "")
            ])
    elif ref_name == "card_types":
        ws.append(["Вид карты", "Наименование для печати"])
        for ct in get_card_types():
            ws.append([ct.get("name", ""), ct.get("print_name", "")])
    elif ref_name == "owners":
        ws.append(["ФИО владельца"])
        for o in get_owners():
            ws.append([o.get("full_name", "")])
    elif ref_name == "applicants":
        ws.append(["ФИО заявителя"])
        for a in get_applicants():
            ws.append([a.get("full_name", "")])
    elif ref_name == "organizations":
        ws.append(["Наименование организации"])
        for o in get_organizations():
            ws.append([o.get("name", "")])
    elif ref_name == "mfcs":
        ws.append(["Код", "Наименование"])
        for m in get_mfcs():
            ws.append([m.get("code", ""), m.get("name", "")])
    else:
        flash("Неизвестный справочник", "danger")
        return redirect(url_for("index"))
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"export_{ref_name}.xlsx", as_attachment=True)


# ============== RUN ==============
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
